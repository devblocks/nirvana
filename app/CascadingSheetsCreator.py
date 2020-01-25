import pandas as pd
import openpyxl as oxl
import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows

#report = xlr.Workbook('groupings.xlsx')
#ws1 = report.add_worksheet('Sheet 1')

class CascadingSheetsCreator(object):
	#TODO - Add hyperlink to A1 to go back to previous page
	#TODO - Add Total at the bottom of each sheet
	#TODO - Globally sort rows on each sheet based
		#based on criteria in the proposal

	headers_list = [] # This should be passed in from GUI, so user
					  # specify desired branches

	def __init__(self):
		#Before this class/function comes into action
		#	have every necessary field already generated
		self.wb = oxl.Workbook()

	def check_headers(self,config):
		try: 
			print(config['headers'])
		except: 
			raise Exception('There must be a headers key in the config')

        # Make sure that each of the config values are lists and not strings
		[config.update({el:[config[el]]}) for el in config.keys() if str(config[el]).find('[') < 0]

        #Make sure that the length of the headers is same for all activities that are labeled
		try:
			elem_len = len(config['headers'])
			for el in config.keys():
				if elem_len != len(config[el]) and config[el] != ['Unlabeled']:
					raise Exception('All labeled keys within the config file must have lists of the same length')
		except:
			pass

				#If non-constant length branching is implemented:
				# el_len = len(config[el])
				# if elem_len > el_len:
				# 	add_space = elem_len - el_len
				# 	for num in range(add_space):
				# 		config[el].append('')
				# if elem_len < el_len:
				# 	del_space = el_len - elem_len
				# 	for num in range(del_space):
				# 		config[el].remove(config[el][len(config[el])-1])

	def create_pivot_table(self,df,headers_list,create_excel=False):
					
		#Create pivot table from hierarchy in GUI -> headers_list
	
		pt_dict = {}

		#Save the pivot tables to an excel sheet
		if create_excel:
			writer = pd.ExcelWriter('report_pivot_table.xlsx', engine='openpyxl') 
		for count,headers in enumerate(headers_list):
			pt = pd.pivot_table(df,index=headers_list[:count+1], aggfunc=np.sum,fill_value="", margins=True, margins_name='Total')
			pt_dict.update({'level'+str(count+1): pt})
			if create_excel:
				pt_dict['level'+str(count+1).to_excel(writer,sheet_name='level'+str(count+1))]
		return pt_dict

	def generate(self,startdf,headers_list):
		# try:
		# 	check_dataframe(startdf,headers_list)
		# except: 
		# 	#send to Redis that df isn't the right format for pushing to
		# 	#the GUI for the user
		pt_dict = self.create_pivot_table(startdf,headers_list)
		self.create_top_sheet(pt_dict['level1'],self.wb)
		if len(headers_list) > 1:
			self.generate_recursive(pt_dict,self.wb,level=1,chain='Total')
		else:
			self.wb.save(filename='CascadingReport.xlsx')
		#Since there has to be a workbook passed into this function, that 
		#	workbook must already have a first sheet. That sheet should be 
		#	titled 'Summary' or 'Category'. This function should start on
		# level 1 and chain = ''.

	def pt_traverser(self,df,dive_list):
	    for count,turns in enumerate(dive_list):
	        df = getattr(df,'loc')[turns]
	    return df

	def create_top_sheet(self,df,wb):
	    #pass in df as "pt_dict['level1']"
	    sheet1 = wb['Sheet']
	    sheet1.title = 'Total'
	    df2 = df.reset_index()
	    for row in dataframe_to_rows(df2):
	        row.remove(row[0])
	        sheet1.append(row) 

	def generate_recursive(self,pt_dict,wb,level=1,chain='Total',dive_list=[]):
	    #TODO - give callback at A1 to origin sheet
	    rows, elem_list = [],[]
	    for row in self.pt_traverser(pt_dict['level'+str(level)],dive_list).iterrows(): # get the index rows of the dataframe
	        rows.append(row[0])
	    #for level 1, rows.remove('Total')
	    if level == 1:
	        rows.remove('Total')

	    for count,items in enumerate(rows): #run through rows of the first column
	        if items not in elem_list: #if there is a new item...
	            dive_list.append(items)
	            new_sheet = self.pt_traverser(pt_dict['level'+str(level+1)],dive_list) # get the dataframe for the next level
	            new_sheet.reset_index(inplace=True) # prep the dataframe for input to the sheet
	            print('Creating '+str(chain)+'-'+str(items))
	            ws2 = wb.create_sheet(title=chain+'-'+items) # create the new sheet
	            for r in dataframe_to_rows(new_sheet): # Log the dataframe to a sheet
	                r.remove(r[0])
	                ws2.append(r)
	            #Add total to bottom of sheet
	            sheet1 = wb[str(chain)]
	            sheet1.cell(row=count+2,column=1).hyperlink = '#'+str(chain)+'-'+str(items)
	            sheet1.cell(row=count+2,column=1).hyperlink.display = str(items)
	            elem_list.append(items) # update the list of unique rows
	        if pt_dict.has_key('level'+str(level+2)):
	            self.generate_recursive(pt_dict,wb,level+1,chain+'-'+items,dive_list)
	        dive_list.remove(items)
	    wb.save(filename='CascadingReport.xlsx')