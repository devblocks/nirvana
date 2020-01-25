from openpyxl import load_workbook

class DataFormatter(object):

    def __init__(self):
        pass

    def nirvana_format(self, excelFile):

        # Load the report's excel file
        wb = load_workbook(excelFile)
        ws = wb.active  # create an object for the first sheet in the file

        # Format a column, one row at a time
        maxRow = ws.max_row
        for row in range(1, maxRow+1):
            _cell = ws.cell(row=row, column=6)  # Select 6th column (Activity Cost)
            _cell.number_format = '$#,##0.00'

            _cell = ws.cell(row=row, column=7)  # Activity Revenue
            _cell.number_format = '$#,##0.00'

            _cell = ws.cell(row=row, column=10)  # Closed-Won Value
            _cell.number_format = '$#,##0.00'

            _cell = ws.cell(row=row, column=12)  # Closed-Lost Value
            _cell.number_format = '$#,##0.00'

            _cell = ws.cell(row=row, column=14)  # Open Value
            _cell.number_format = '$#,##0.00'

        wb.save(filename='CascadingReport.xlsx')

if __name__ == "__main__":
    ff = DataFormatter()
    ff.nirvana_format('../CascadingReport.xlsx')